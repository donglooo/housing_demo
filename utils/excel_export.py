import pandas as pd
import re
import json
import os
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
import yaml

class ExcelExporter:
    def _sanitize_df(self, df):
        """Removes illegal characters (control chars) that break Excel export."""
        if df.empty: return df
        # Regex for illegal control chars: ASCII 0-31 excluding 9 (Tab), 10 (LF), 13 (CR)
        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
        
        def clean_text(text):
            if isinstance(text, str):
                return ILLEGAL_CHARACTERS_RE.sub('', text)
            return text
            
        df_clean = df.copy()
        
        # 1. Clean Data (Object cols)
        for col in df_clean.select_dtypes(include=[object]):
            df_clean[col] = df_clean[col].apply(clean_text)
            
        # 2. Clean Index
        if isinstance(df_clean.index, pd.MultiIndex):
            new_levels = []
            for level in df_clean.index.levels:
                if level.dtype == object:
                    new_levels.append(level.map(clean_text))
                else:
                    new_levels.append(level)
            df_clean.index = df_clean.index.set_levels(new_levels, level=range(len(new_levels)))
        elif df_clean.index.dtype == object:
            # Handle Mixed Index types (Tuples + Strings) which crash pandas if assigned back
            try:
                df_clean.index = df_clean.index.map(clean_text)
            except TypeError:
                # Fallback: If pandas fails to handle mixed tuple/str assignment
                # We skip index sanitization for this edge case to prevent crash.
                pass            
        # 3. Clean Columns
        if isinstance(df_clean.columns, pd.MultiIndex):
            new_levels = []
            for level in df_clean.columns.levels:
                if level.dtype == object:
                    new_levels.append(level.map(clean_text))
                else:
                    new_levels.append(level)
            df_clean.columns = df_clean.columns.set_levels(new_levels, level=range(len(new_levels)))
        elif df_clean.columns.dtype == object:
            df_clean.columns = df_clean.columns.map(clean_text)
            
        return df_clean

    def __init__(self):
        # Load Config
        self.style_config = {}
        try:
            with open("config/config.yaml", "r") as f:
                config = yaml.safe_load(f)
                self.style_config = config.get("excel_style", {})
        except:
            pass

        # Defaults
        self.FONT_NAME = self.style_config.get("font_name", "Yu Gothic UI")
        
        # Colors (Strip '#' if present for openpyxl)
        def clean_color(c): return str(c).replace("#", "")
        
        self.HEADER_BG_COLOR = clean_color(self.style_config.get("header_bg_color", "0070C0"))
        self.HEADER_FONT_COLOR = clean_color(self.style_config.get("header_font_color", "FFFFFF"))
        
        self.GRAY_BG = "E7E6E6"      # Keep Fixed for Index
        self.WHITE_BG = "FFFFFF"
        
        # Heatmap Colors
        self.HEATMAP_MIN = clean_color(self.style_config.get("heatmap_min_color", "F8696B"))
        self.HEATMAP_MID = clean_color(self.style_config.get("heatmap_mid_color", "FFFFFF"))
        self.HEATMAP_MAX = clean_color(self.style_config.get("heatmap_max_color", "FFC000"))
        
        # Styles
        self.header_fill = PatternFill(start_color=self.HEADER_BG_COLOR, end_color=self.HEADER_BG_COLOR, fill_type="solid")
        self.header_font = Font(name=self.FONT_NAME, color=self.HEADER_FONT_COLOR, bold=True)
        
        self.gray_fill = PatternFill(start_color=self.GRAY_BG, end_color=self.GRAY_BG, fill_type="solid")
        self.white_fill = PatternFill(start_color=self.WHITE_BG, end_color=self.WHITE_BG, fill_type="solid")
        
        # Data Fonts
        self.data_font = Font(name=self.FONT_NAME, color="000000")
        self.total_font = Font(name=self.FONT_NAME, color="000000", bold=True)
        
        self.data_border = Border(
            left=Side(style='thin', color="000000"), 
            right=Side(style='thin', color="000000"), 
            top=Side(style='thin', color="000000"), 
            bottom=Side(style='thin', color="000000")
        )
        self.align_center = Alignment(horizontal='center', vertical='center')
        self.align_right = Alignment(horizontal='right', vertical='center')
        self.align_left = Alignment(horizontal='left', vertical='center')

    def write_table(self, writer, df, sheet_name, start_row, start_col, num_fmt='0'):
        """
        Writes a single dataframe to the sheet with styling.
        start_row, start_col are 1-based.
        Returns (max_row, max_col) of the written area.
        """
        # Sanitize DataFrame (remove illegal characters)
        df = self._sanitize_df(df)
        
        # Write to Excel
        # startrow/startcol in to_excel are 0-based.
        df.to_excel(writer, sheet_name=sheet_name, startrow=start_row-1, startcol=start_col-1)
        
        ws = writer.sheets[sheet_name]
        
        n_rows, n_cols = df.shape
        
        # Calculate Dimensions including Headers and Index
        header_levels = df.columns.nlevels
        index_levels = df.index.nlevels
        
        total_cols = n_cols + index_levels 
        total_rows = n_rows + header_levels
        
        min_row = start_row
        max_row = min_row + total_rows - 1
        min_col = start_col
        max_col = min_col + total_cols - 1
        
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=r, column=c)
                
                # Default Border
                cell.border = self.data_border
                
                # Check Labels for Total/Indices
                # Header Area: Rows [min_row, min_row + header_levels - 1]
                is_header_row = (r < min_row + header_levels)
                
                # Index Area: Cols [min_col, min_col + index_levels - 1]
                # But headers also span index cols.
                is_index_col = (c < min_col + index_levels)
                
                # Heuristic for Total Row/Col
                # If the cell value is 'Total' or '總計', or if it's in the last row/col and we know it's a total
                # We simply check if the value suggests it is a header type
                val_str = str(cell.value) if cell.value else ""
                is_total_label = "Total" in val_str or "總計" in val_str
                
                # Styling Logic
                if is_header_row:
                    cell.fill = self.header_fill
                    cell.font = self.header_font
                    cell.alignment = self.align_center
                elif is_index_col:
                    cell.fill = self.gray_fill
                    cell.font = self.data_font
                    cell.number_format = '@' # Text
                    cell.alignment = self.align_center
                    # If it's the Total Label
                    if is_total_label:
                         cell.font = self.total_font
                else:
                    # Data Area
                    # Check if it falls in the "Total" row or column
                    # We can infer Total Row if the Index Col of this row is "Total"
                    index_val = ws.cell(row=r, column=min_col).value
                    col_header_val = ws.cell(row=min_row, column=c).value
                    
                    is_total_row = str(index_val) in ["Total", "總計"]
                    is_total_col = str(col_header_val) in ["Total", "總計"]
                    
                    if is_total_row or is_total_col:
                        cell.fill = self.gray_fill
                        cell.font = self.total_font
                    else:
                        cell.fill = self.white_fill
                        cell.font = self.data_font
                        
                    cell.number_format = num_fmt
                    cell.alignment = self.align_right

        # Apply Data Bars to Numeric Data Range (excluding headers and index col)
        # Range: (min_row + header_levels) to max_row, (min_col + index_levels) to max_col
        data_min_row = min_row + header_levels
        data_min_col = min_col + index_levels
        data_max_row = max_row
        data_max_col = max_col
        
        # Helper to detect Total label (Handles Index/MultiIndex)
        def is_total(label):
            # Check list/tuple (MultiIndex)
            if isinstance(label, (list, tuple)):
                # If ANY level says Total/總計, treat as Total
                return any(str(x).strip() in ["總計", "Total"] for x in label)
            
            # Check single string/value
            s_label = str(label).strip()
            return s_label in ["總計", "Total"]

        # Exclude Total Row if present
        # Check LAST item
        if not df.empty and is_total(df.index[-1]):
            data_max_row -= 1
            
        # Exclude Total Column if present
        # Check LAST item
        if not df.empty and is_total(df.columns[-1]):
            data_max_col -= 1
        
        if data_max_row >= data_min_row and data_max_col >= data_min_col:
            start_cell = f"{get_column_letter(data_min_col)}{data_min_row}"
            end_cell = f"{get_column_letter(data_max_col)}{data_max_row}"
            range_ref = f"{start_cell}:{end_cell}"

            # FIX: Add Zero Rule FIRST to ensure it takes precedence (Priority 1)
            # 1. Zero Value Rule: If 0, paint White
            zero_rule = CellIsRule(
                operator='equal', 
                formula=['0'], 
                stopIfTrue=True, 
                fill=self.white_fill
            )
            zero_rule.priority = 1 # High Priority
            ws.conditional_formatting.add(range_ref, zero_rule)

            # 2. Dynamic Colors (Heatmap) - Lower Priority
            heatmap_rule = ColorScaleRule(
                start_type='min', start_color=self.HEATMAP_MIN,
                mid_type='num', mid_value=0, mid_color=self.HEATMAP_MID,
                end_type='max', end_color=self.HEATMAP_MAX
            )
            # heatmap_rule.priority = 2 # ColorScaleRule in openpyxl might not support simple priority setting easily, 
            # but being added SECOND usually puts it lower in the stack or simply below the StopIfTrue rule.
            # IN Excel: Top Rule Wins. ensure Zero is Top.
            ws.conditional_formatting.add(range_ref, heatmap_rule)
        
        # Return bounds AND index information for later column sizing
        # Identifying index columns range: [min_col, min_col + index_levels - 1]
        index_cols = list(range(min_col, min_col + index_levels))
        # Data columns range: [min_col + index_levels, max_col]
        data_cols = list(range(min_col + index_levels, max_col + 1))
        
        return max_row, max_col, index_cols, data_cols

    def load_users(self):
        try:
            with open("config/config.yaml", "r") as f:
                return yaml.safe_load(f)["users"]
        except:
            return ["User 1", "User 2"]

    def _render_sidebar(self, ws, users):
        """Renders Vertical Sidebar: User Sign-off (No Legend)"""
        # --- Headers at Row 2 ---
        # B2: "人員", C2: "到此一遊"
        c_person = ws.cell(row=2, column=2, value="人員")
        c_person.font = self.header_font
        c_person.fill = self.header_fill 
        c_person.alignment = self.align_center
        c_person.border = self.data_border

        c_visited = ws.cell(row=2, column=3, value="到此一遊")
        c_visited.font = self.header_font
        c_visited.fill = self.header_fill
        c_visited.alignment = self.align_center
        c_visited.border = self.data_border
        
        # Validation Options
        options = '"➖,✅,💬,💡"'
        dv = DataValidation(type="list", formula1=options, allow_blank=True)
        ws.add_data_validation(dv)
        
        # --- User List (Start B3) ---
        start_row = 3
        for i, user in enumerate(users):
            r = start_row + i
            
            # Name (B)
            c_name = ws.cell(row=r, column=2, value=user)
            c_name.font = self.header_font
            c_name.fill = self.header_fill # Blue
            c_name.alignment = self.align_center
            c_name.border = self.data_border
            
            # Status Dropdown (C)
            c_status = ws.cell(row=r, column=3, value="➖")
            c_status.alignment = self.align_center
            c_status.border = self.data_border
            dv.add(c_status)

    def create_index_sheet(self, writer, data_dict, source_name, gen_time, toc_metadata=None):
        """
        Creates a Homepage/Index sheet with links to all other sheets.
        toc_metadata: List of dicts {'sheet_name': str, 'description': str, 'index': str} (Optional)
        """
        sheet_name = "首頁"
        if sheet_name not in writer.book.sheetnames:
            writer.book.create_sheet(sheet_name, 0) # Create at index 0
        
        ws = writer.book[sheet_name]
        ws.title = sheet_name
        
        users = self.load_users()
        
        # --- Constants ---
        FONT_NAME = self.FONT_NAME
        COLOR_DARK_BLUE = "0F243E"
        # COLOR_LIGHT_BLUE = "538DD5" 
        # Use Header Config for Index Banner? Maybe keep fixed for Index Page aesthetics or use Config?
        # Let's keep Index Page standard for now, but use Config Font
        
        COLOR_LIGHT_BLUE = "538DD5"
        COLOR_LIGHT_GRAY = "F2F2F2"
        COLOR_WHITE = "FFFFFF"
        COLOR_BLACK = "000000"
        
        # Fills
        fill_base = PatternFill(start_color=COLOR_DARK_BLUE, end_color=COLOR_DARK_BLUE, fill_type="solid")
        fill_banner = PatternFill(start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid")
        fill_content = PatternFill(start_color=COLOR_LIGHT_GRAY, end_color=COLOR_LIGHT_GRAY, fill_type="solid")
        
        # Fonts
        font_title = Font(name=FONT_NAME, size=20, bold=True, color=COLOR_WHITE)
        font_white = Font(name=FONT_NAME, color=COLOR_WHITE, bold=True)
        font_black_bold = Font(name=FONT_NAME, color=COLOR_BLACK, bold=True)
        font_black = Font(name=FONT_NAME, color=COLOR_BLACK)
        font_link = Font(name=FONT_NAME, color="0563C1", underline="single")
        
        # 1. Base Fill
        for r in range(1, 101):
            for c in range(1, 40): 
                ws.cell(row=r, column=c).fill = fill_base
                
        # 2. Banner Fill (B2:V4)
        for r in range(2, 5):
            for c in range(2, 23):
                ws.cell(row=r, column=c).fill = fill_banner
                
        # 3. Content Fill (B5:V80)
        for r in range(5, 81):
            for c in range(2, 23):
                ws.cell(row=r, column=c).fill = fill_content

        # --- Layout ---
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 10

        if toc_metadata:
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 65
        
        ws.row_dimensions[3].height = 30 
        
        # --- Content ---
        
        # Title at C3
        cell_title = ws.cell(row=3, column=3, value="資料匯出報表目錄")
        cell_title.font = font_title
        cell_title.alignment = self.align_left
        
        # Metadata
        # Source at N3
        cell_source = ws.cell(row=3, column=14, value=f"資料源：{source_name}")
        cell_source.font = font_white
        cell_source.alignment = self.align_left
        
        # Date at T3
        cell_date = ws.cell(row=3, column=20, value=gen_time) 
        cell_date.font = font_white
        cell_date.alignment = self.align_left
        
        # Headers at Row 6
        start_row = 6
        
        if toc_metadata:
             # Custom Headers
             headers = ["序號", "表索引", "描述", "連結"] + users
             start_col = 3
        else:
             # Legacy Headers
             headers = ["序號", "項目", "連結"] + users
             start_col = 3
        
        medium_border = Border(bottom=Side(style='medium', color="000000"))
        
        for i, h in enumerate(headers):
            cell = ws.cell(row=start_row, column=start_col + i, value=h)
            cell.font = font_black_bold
            cell.border = medium_border
            cell.alignment = self.align_center
            
        # Data List at Row 7
        row = 7
        
        # Determine iterator
        if toc_metadata:
            # Ensure we only list sheets that actually exist in data_dict
            items_to_list = [m for m in toc_metadata if m['sheet_name'] in data_dict]
            
            for idx, item in enumerate(items_to_list, 1):
                name = item['sheet_name']
                desc = item.get('description', '')
                idx_label = item.get('index', name)
                
                # Seq
                c_seq = ws.cell(row=row, column=start_col, value=idx)
                c_seq.font = font_black
                c_seq.alignment = self.align_center
                
                # Index (表索引)
                c_idx = ws.cell(row=row, column=start_col + 1, value=idx_label)
                c_idx.font = font_black
                c_idx.alignment = self.align_left
                
                # Description (描述)
                c_desc = ws.cell(row=row, column=start_col + 2, value=desc)
                c_desc.font = font_black
                c_desc.alignment = self.align_left
                
                # Link
                c_link = ws.cell(row=row, column=start_col + 3, value="查看")
                c_link.hyperlink = f"#'{name}'!A1"
                c_link.font = font_link
                c_link.alignment = self.align_center
                
                # User Status
                safe_sheet_name = name.replace("'", "''") 
                for u_i, user in enumerate(users):
                     target_row = 3 + u_i
                     cell_status = ws.cell(row=row, column=start_col + 4 + u_i)
                     cell_status.value = f"='{safe_sheet_name}'!C{target_row}"
                     cell_status.alignment = self.align_center
                     cell_status.font = font_black
                
                row += 1
                
        else:
            # Legacy Logic
            for idx, name in enumerate(data_dict.keys(), 1):
                # Seq
                c_seq = ws.cell(row=row, column=start_col, value=idx)
                c_seq.font = font_black
                c_seq.alignment = self.align_center
                
                # Item
                c_name = ws.cell(row=row, column=start_col + 1, value=name)
                c_name.font = font_black
                c_name.alignment = self.align_left
                
                # Link
                c_link = ws.cell(row=row, column=start_col + 2, value="查看")
                c_link.hyperlink = f"#'{name}'!A1"
                c_link.font = font_link
                c_link.alignment = self.align_center
                
                # User Status
                safe_sheet_name = name.replace("'", "''") 
                for u_i, user in enumerate(users):
                     target_row = 3 + u_i
                     cell_status = ws.cell(row=row, column=start_col + 3 + u_i)
                     cell_status.value = f"='{safe_sheet_name}'!C{target_row}"
                     cell_status.alignment = self.align_center
                     cell_status.font = font_black
                
                row += 1
            
        ws.sheet_view.showGridLines = False

    def generate_report(self, data_dict, output, source_file="Unknown", gen_time="Unknown", toc_metadata=None):
        writer = pd.ExcelWriter(output, engine='openpyxl')
        users = self.load_users()
        
        # Determine Sheet Order
        if toc_metadata:
             # Use metadata order, finding matching data
             ordered_items = []
             for m in toc_metadata:
                 s_name = m['sheet_name']
                 if s_name in data_dict:
                     ordered_items.append((s_name, data_dict[s_name]))
        else:
             # Default dict order -> Convert to list for indexing
             ordered_items = list(data_dict.items())
        
        # 1. Generate Content Sheets
        for i, (sheet_name, sheet_data) in enumerate(ordered_items):
            # Track Column Types for Widths
            index_cols_set = set()
            data_cols_set = set()
            
            # Standard Params
            current_row = 2 
            start_col = 5   # Col E
            
            if isinstance(sheet_data, dict) and sheet_data.get('mode') == 'horizontal_split':
                # --- Horizontal Block Mode ---
                blocks = sheet_data.get('blocks', [])
                current_block_start_col = 5 # Start E
                
                # Check for empty (Prevent crash if no data)
                if not blocks: continue
                
                # Lookup Description
                current_desc = ""
                if toc_metadata:
                     current_desc = next((m.get('description', '') for m in toc_metadata if m['sheet_name'] == sheet_name), "")
                
                for block in blocks:
                    block_tables = block.get('tables', [])
                    block_year = block.get('year', '')
                    
                    # Reset Row for each Block (Top Aligned)
                    current_row = 3 # Start Data at Row 3 (Row 2 reserved for Header)
                    block_max_col = current_block_start_col
                    
                    for i_table, table_info in enumerate(block_tables):
                        df = table_info['df']
                        fmt = table_info['fmt']
                        
                        # Write Table
                        # Updated write_table returns col metadata
                        end_row, end_col, idx_cols, d_cols = self.write_table(writer, df, sheet_name, current_row, current_block_start_col, fmt)
                        
                        index_cols_set.update(idx_cols)
                        data_cols_set.update(d_cols)
                        
                        # Write Year Header (After first write ensures sheet exists)
                        if i_table == 0:
                            ws = writer.sheets[sheet_name]
                            year_cell = ws.cell(row=2, column=current_block_start_col)
                            
                            # Checks if block_year is a number (e.g. 2024), if so append '年', else keep as is (e.g. YOY)
                            disp_year = str(block_year)
                            if disp_year.isdigit() or (disp_year.replace('.','',1).isdigit()):
                                year_cell.value = f"{disp_year}年"
                            else:
                                year_cell.value = disp_year
                                
                            year_cell.font = Font(bold=True, size=14, color="000000")
                            year_cell.alignment = Alignment(horizontal='left')
                            
                            # Write Description (Next to Year)
                            if current_desc:
                                desc_cell = ws.cell(row=2, column=current_block_start_col + 1, value=current_desc)
                                desc_cell.font = Font(name=self.FONT_NAME, bold=True, color="000000")
                                desc_cell.alignment = Alignment(horizontal='left')
                        
                        # Stack Vertically within block
                        current_row = end_row + 2
                        block_max_col = max(block_max_col, end_col)
                    
                    # Update Start Col for NEXT Block
                    # Previous Max + Gap (1 col) -> Start at Max + 2
                    current_block_start_col = block_max_col + 2
                    
            else:
                # --- Standard Mode ---
                # sheet_data is List[Tables] or {'mode': 'standard', 'tables': ...}
                tables = []
                if isinstance(sheet_data, list):
                    tables = sheet_data
                elif isinstance(sheet_data, dict):
                    tables = sheet_data.get('tables', [])
                
                if not tables: continue

                # Lookup Description
                current_desc = ""
                if toc_metadata:
                     current_desc = next((m.get('description', '') for m in toc_metadata if m['sheet_name'] == sheet_name), "")
                    
                for i_table, table_info in enumerate(tables):
                    df = table_info['df']
                    fmt = table_info['fmt']
                    
                    # Write table
                    end_row, end_col, idx_cols, d_cols = self.write_table(writer, df, sheet_name, current_row, start_col, fmt)
                    
                    index_cols_set.update(idx_cols)
                    data_cols_set.update(d_cols)
                    
                    # Write Description (Standard Mode - Row 1)
                    if i_table == 0 and current_desc:
                        ws = writer.sheets[sheet_name]
                        desc_cell = ws.cell(row=1, column=start_col + 1, value=current_desc)
                        desc_cell.font = Font(name=self.FONT_NAME, bold=True, color="000000")
                        desc_cell.alignment = Alignment(horizontal='left')
                    
                    # Update for next table
                    current_row = end_row + 2 
            
            # Post-processing & Sidebar (ONCE per sheet)
            ws = writer.sheets[sheet_name]
            self._render_sidebar(ws, users)
            
            ws.sheet_view.showGridLines = False
            
            # Navigation Links
            link_font = Font(color="0563C1", underline="single", bold=True)
            
            # A1: Home
            ws["A1"] = "🏠"
            ws["A1"].hyperlink = "#'首頁'!C3"
            ws["A1"].font = link_font
            
            # B1: Prev (if exists)
            if i > 0:
                prev_name = ordered_items[i-1][0]
                safe_prev = prev_name.replace("'", "''")
                ws["A2"] = "⬅"
                ws["A2"].hyperlink = f"#'{safe_prev}'!A2"
                ws["A2"].font = link_font
            
            # C1: Next (if exists)
            if i < len(ordered_items) - 1:
                next_name = ordered_items[i+1][0]
                safe_next = next_name.replace("'", "''")
                ws["A3"] = "➡"
                ws["A3"].hyperlink = f"#'{safe_next}'!A3"
                ws["A3"].font = link_font
            
            # Column Widths Implementation
            first_col_w = self.style_config.get("first_col_width", 25)
            data_col_w = self.style_config.get("data_col_width", 12)
            
            # 1. Apply tracked widths
            for c_idx in index_cols_set:
                col_letter = get_column_letter(c_idx)
                ws.column_dimensions[col_letter].width = first_col_w
                
            for c_idx in data_cols_set:
                col_letter = get_column_letter(c_idx)
                ws.column_dimensions[col_letter].width = data_col_w
            
            # 2. Fixed Sidebar Widths and Nav Widths
            # B and C are used for nav buttons now, so maybe give them some space if not used by data?
            # Data starts at E (5), so A, B, C, D are margin/metadata.
            ws.column_dimensions['A'].width = 3 # Home
            ws.column_dimensions['B'].width = 8 # Prev
            ws.column_dimensions['C'].width = 12 # Next
            ws.column_dimensions['D'].width = 3  # Spacer
                
        # 2. Generate Index Sheet
        self.create_index_sheet(writer, data_dict, source_file, gen_time, toc_metadata)
        
        writer.close()
